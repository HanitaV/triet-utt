import docx
import openpyxl
import os

# Paths
docx_path = r"c:\Users\eleven\triet-utt\subjects\utt\log\Đề cương ôn tập Đại cương Logistics và chuỗi cung ứng.docx"
xlsx_path = r"c:\Users\eleven\triet-utt\subjects\utt\log\Đáp án đề cương Đại cương Logistics và chuỗi cung ứng.xlsx"
output_dir = r"c:\Users\eleven\triet-utt\subjects\utt\log"

# Extract from docx
docx_content = []
doc = docx.Document(docx_path)
for para in doc.paragraphs:
    text = para.text.strip()
    if text:
        docx_content.append(text)

# Save docx content to a text file
with open(os.path.join(output_dir, "docx_content.txt"), "w", encoding="utf-8") as f:
    f.write("\n".join(docx_content))

print(f"Saved {len(docx_content)} paragraphs from docx")

# Extract from xlsx
xlsx_content = []
wb = openpyxl.load_workbook(xlsx_path)
for sheet_name in wb.sheetnames:
    xlsx_content.append(f"--- Sheet: {sheet_name} ---")
    ws = wb[sheet_name]
    for row in ws.iter_rows(values_only=True):
        row_data = [str(cell) if cell is not None else "" for cell in row]
        if any(row_data):
            xlsx_content.append(" | ".join(row_data))

# Save xlsx content to a text file
with open(os.path.join(output_dir, "xlsx_content.txt"), "w", encoding="utf-8") as f:
    f.write("\n".join(xlsx_content))

print(f"Saved {len(xlsx_content)} rows from xlsx")
